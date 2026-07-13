"""PDF (ReportLab) and Excel (openpyxl) report generation."""
import io
from datetime import datetime

from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

HEADER_BLUE = "1D4ED8"


def excel_response(title, headers, rows, filename):
    """Build an .xlsx file from headers + row tuples and return it."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(size=13, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)
    cell.alignment = Alignment(horizontal="center")

    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=header)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DBEAFE")

    for r, row in enumerate(rows, start=3):
        for col, value in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=value)

    for col in range(1, len(headers) + 1):
        width = max(
            [len(str(headers[col - 1]))] +
            [len(str(row[col - 1])) for row in rows if len(row) >= col]
        ) if rows else len(str(headers[col - 1]))
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 3, 10), 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"{filename}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def pdf_response(title, headers, rows, filename, subtitle=None, meta=None):
    """Build a tabular PDF report and return it."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], fontSize=15,
        textColor=colors.HexColor("#" + HEADER_BLUE), spaceAfter=2,
    )
    elements = [Paragraph(title, title_style)]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    if meta:
        for line in meta:
            elements.append(Paragraph(line, styles["Normal"]))
    elements.append(Paragraph(
        f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ParagraphStyle("Meta", parent=styles["Normal"], textColor=colors.grey, fontSize=8),
    ))
    elements.append(Spacer(1, 6 * mm))

    body_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8.5)
    data = [headers] + [
        [Paragraph(str(v), body_style) if isinstance(v, str) and len(str(v)) > 28 else v
         for v in row]
        for row in rows
    ]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + HEADER_BLUE)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"{filename}.pdf",
        mimetype="application/pdf",
    )
